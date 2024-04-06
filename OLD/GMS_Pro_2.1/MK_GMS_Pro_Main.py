from IPython import get_ipython
get_ipython().run_line_magic('clear','')
get_ipython().run_line_magic('reset', '-f')
PERIOD = 0
GMS_NAME      = 'Pro021'
GMS_VERSION   = 'GMS_Pro_2.1'
SETUP_FILE    = 'MK_GMS_Pro-Setup.npz'
MAIN_DIR      = 'C:/Users/Ich/Documents/!1 Vorlesungen/!PlanSpiele/'
SUB_DIR       = '!GMS_Versions'
SCEN_DIR      = 'Szenario'
GMS_DIR       = '!GMS_Ordner'
GMS_directory = MAIN_DIR + GMS_DIR + '/' + GMS_NAME+'/'
print('Planspiel-Verzeichnis')
print(GMS_directory)
import sys
version_dir = MAIN_DIR + SUB_DIR + '/' + GMS_VERSION + '/'
print(version_dir)
sys.path.append(version_dir)
import time
from openpyxl import load_workbook
import xlwings            as xw
import numpy              as np
import MK_GMS_Pro_Modules as mod
np.set_printoptions(precision=4, suppress=True)
start_time = time.time()
with np.load(GMS_directory + SETUP_FILE) as setup:
    gms_numbers = setup['gms_numbers']
    gms_files   = setup['gms_files']
del(setup)
NUM_COMPANIES  = gms_numbers[0]
num_companies0 = NUM_COMPANIES
NUM_PERIODS    = gms_numbers[1]
OFFSET         = gms_numbers[2]
NUM_MARKETS    = gms_numbers[3]
MARKET_0       = gms_numbers[4]
MARKET_1       = gms_numbers[5]
IDEAL_RD       = gms_numbers[8]
MARKET_2       = gms_numbers[6]
MARKET_3       = gms_numbers[7]
NUM_CELLS      = gms_numbers[9]
UL_CELLS       = gms_numbers[10]
COST_INDUSTRY_REPORT = gms_numbers[11]
COST_MARKET_REPORT   = gms_numbers[12]
del(gms_numbers)
scenario_dir  = version_dir + SCEN_DIR + '/'
decision_dir  = GMS_directory + 'Decisions/'
company_dir = []
for co in range(NUM_COMPANIES):
    company_dir.append(GMS_directory+f'U{co+1:0>2d}/')
SCENARIO_FILE    = gms_files[0]
COMPANY_FILE     = gms_files[1]
INFO_FILE        = gms_files[2]
info_period_file = f'MK_GMS_Pro-Informationen-P{PERIOD:0>2d}.xlsx'
del(gms_files)
decision_template = 'Entscheidungen-P00-U01.xlsx'
decision_files = []
decision_files_new = []
report_files_old = []
report_files = []
info_files = []
for co in range(NUM_COMPANIES):
    decision_files.append(
        f'Entscheidungen-P{PERIOD:0>2d}-U{co+1:0>2d}.xlsx')
    decision_files_new.append(
        f'Entscheidungen-P{PERIOD+1:0>2d}-U{co+1:0>2d}.xlsx')
    report_files_old.append(
        f'Berichte-P{PERIOD-1:0>2d}-U{co+1:0>2d}.xlsx')
    report_files.append(
        f'Berichte-P{PERIOD:0>2d}-U{co+1:0>2d}.xlsx')
    info_files.append(
        f'Informationen-P{PERIOD:0>2d}-U{co+1:0>2d}.xlsx')
if (PERIOD == 0):
    report_files_old = ['Berichte-P(-01)-U01.xlsx']
    NUM_COMPANIES = 1
prod_demand = np.zeros((NUM_MARKETS, num_companies0), dtype=int)
prod_supply = np.zeros((NUM_MARKETS, num_companies0), dtype=int)
sales_cur   = np.zeros((NUM_MARKETS, num_companies0), dtype=int)
del(NUM_MARKETS)
mDec_SOLID_cell  = ('D6' , 'D7' , 'D10', 'D11',
                    'D14', 'D15', 'D20',
                    'D27', 'D19')
mDec_IDEAL_cell  = ('E6' , 'E7' , 'E10', 'E11',
                    'E14', 'E15', 'E20',
                    'E27')
mDec_GESAMT_cell = ('F16', 'F21', 'F22', 'F26')
mDec_SOLID       = np.zeros((len(mDec_SOLID_cell) , NUM_COMPANIES), dtype=int)
mDec_IDEAL       = np.zeros((len(mDec_IDEAL_cell) , NUM_COMPANIES), dtype=int)
mDec_GESAMT      = np.zeros((len(mDec_GESAMT_cell), NUM_COMPANIES), dtype=int)
pDec_SOLID_cell  = ('J6' , 'J9' , 'J10')
pDec_IDEAL_cell  = ('K6' , 'K9' , 'K11')
pAll_TA_neu_cell = ('J14', 'J21')
pAll_TA_alt_cell = ('K14', 'K21')
pDec_HR_cell     = ('L26', 'L27',
                    'L28', 'L29')
pDec_SOLID       = np.zeros((len(pDec_SOLID_cell) , NUM_COMPANIES), dtype=int)
pDec_IDEAL       = np.zeros((len(pDec_IDEAL_cell) , NUM_COMPANIES), dtype=int)
pDec_HR          = np.zeros((len(pDec_HR_cell)    , NUM_COMPANIES))
pAll_TA_neu      = np.empty((NUM_CELLS, NUM_COMPANIES), dtype=object)
pAll_TA_alt      = np.empty((NUM_CELLS, NUM_COMPANIES), dtype=object)
pAll_TA_inv      = np.zeros((NUM_CELLS, NUM_COMPANIES), dtype=int)
pAll_TA_sell     = np.zeros((NUM_CELLS, NUM_COMPANIES), dtype=int)
pAll_TA_buy      = np.zeros((NUM_CELLS, NUM_COMPANIES), dtype=int)
fDec_FIN_cell    = ('P6' , 'P7' , 'P8' , 'P11', 'P12')
fDec_FIN  = np.zeros((len(fDec_FIN_cell), NUM_COMPANIES), dtype=int)
print('')
print(f'Simuliere Periode {PERIOD:0>1d}')
szenario = dict(np.load(GMS_directory + SCENARIO_FILE, allow_pickle=True))
with np.load(GMS_directory + COMPANY_FILE, allow_pickle=True) as co_file:
    mDec_SOLID_h  = co_file['mDec_SOLID']
    mDec_IDEAL_h  = co_file['mDec_IDEAL']
    mDec_GESAMT_h = co_file['mDec_GESAMT']
    mMix_SOLID_h  = co_file['mMix_SOLID']
    cSat_SOLID_h  = co_file['cSat_SOLID']
    mMix_IDEAL_h  = co_file['mMix_IDEAL']
    cSat_IDEAL_h  = co_file['cSat_IDEAL']
    pDec_SOLID_h  = co_file['pDec_SOLID']
    pDec_IDEAL_h  = co_file['pDec_IDEAL']
    pRes_TA_h     = co_file['pRes_TA']
    pDec_HR_h     = co_file['pDec_HR']
    pRes_HR_h     = co_file['pRes_HR']
    pRes_costs_h  = co_file['pRes_costs']
    fDec_FIN_h    = co_file['fDec_FIN']
    fRes_COMP_h   = co_file['fRes_COMP']
del(co_file)
info_period_WB  = load_workbook(filename = scenario_dir + INFO_FILE)
info_period_WS  = info_period_WB['Aktuelle Wirtschaftsdaten']
szenario_tuple = ('PreisHStoff', 'PreisMat1', 'PreisMat2', 'LagerMat',
                  'Zinsen', 'Finanzen',
                  'Gehalt', 'LohnNK', 'BedarfSOLID', 'BedarfIDEAL', 'AV_GG',
                  'AnlagenS', 'AnlagenM', 'AnlagenL',
                  'Sondermarkt', 'TransportFE', 'LagerFE')
write_range = (('D6' , 'D10' ),
               ('D12', 'D17'),
               ('D19', 'D24'),
               ('D26', 'D28'),
               ('D33', 'D38'),
               ('D40', 'D42'),
               ('I6',  'I8' ),
               ('I10', 'I18'),
               ('I23', 'I26'),
               ('I28', 'I31'),
               ('I36', 'I38'),
               ('N6',  'N11' ),
               ('N13', 'N18'),
               ('N20', 'N25'),
               ('N29', 'N30'),
               ('N32', 'N41'),
               ('N43', 'N46'))
for ndx in range(len(write_range)):
    mod.write_XLS_range(
        write_range[ndx],
        szenario[szenario_tuple[ndx]][:, PERIOD+OFFSET+1],
        info_period_WS)
write_range = (('E6' , 'E10' ),
               ('E12', 'E17'),
               ('E19', 'E24'),
               ('E26', 'E28'),
               ('E33', 'E38'),
               ('E40', 'E42'),
               ('J6',  'J8' ),
               ('J10', 'J18'),
               ('J23', 'J26'),
               ('J28', 'J31'),
               ('J36', 'J38'),
               ('O6',  'O11' ),
               ('O13', 'O18'),
               ('O20', 'O25'),
               ('O29', 'O30'),
               ('O32', 'O41'),
               ('O43', 'O46'))
for ndx in range(len(write_range)):
    mod.write_XLS_range(
        write_range[ndx],
        szenario[szenario_tuple[ndx]][:, PERIOD+OFFSET],
        info_period_WS)
info_period_WS['O1'].value   = num_companies0
info_period_WS['O2'].value   = PERIOD
info_period_WS['I41'].value  = szenario['AV_BGA'][PERIOD+OFFSET+1]
info_period_WS['J41'].value  = szenario['AV_BGA'][PERIOD+OFFSET]
for per in range(NUM_PERIODS+1):
    if per != PERIOD:
        info_period_WS = info_period_WB[f'KonjunkturAusblick P{per+1:0>2d}']
        info_period_WB.remove(info_period_WS)
del(per)
if PERIOD < MARKET_2:
    info_period_WS = info_period_WB['Marktbericht IDEAL']
    info_period_WB.remove(info_period_WS)
mean_time = time.time()
print('')
print('Setup')
print(round(mean_time - start_time, 1), f'({mean_time - start_time:.1f})')
mean_time0 = mean_time
for co in range(NUM_COMPANIES):
    if (PERIOD == 0):
        decisionWB_read  = load_workbook(
            filename = scenario_dir + decision_files[co],
            data_only=True)
    else:
        decisionWB_read  = load_workbook(
            filename = decision_dir + decision_files[co],
            data_only=True)
    decisionWB_write = load_workbook(
        filename = scenario_dir + decision_template)
    decisionWS_read  = decisionWB_read['Entscheidungen']
    decisionWS_write = decisionWB_write['Entscheidungen']
    cell_values = mod.read_XLS_cells(mDec_SOLID_cell, decisionWS_read)
    cell_values = [x if x is not None else 0 for x in cell_values]
    mod.write_XLS_cells(mDec_SOLID_cell, cell_values, decisionWS_write)
    if (cell_values[7] == 'Ja'):
        cell_values[7] = 1
    else:
        cell_values[7] = 0
    mDec_SOLID[:, co] = np.array(cell_values)
    mDec_SOLID[8, co] = min(mDec_SOLID[8, co],
                            szenario['Sondermarkt'][1, PERIOD+OFFSET])
    if PERIOD<MARKET_1:
        mDec_SOLID[[1, 5, 6]] = 0
    cell_values = mod.read_XLS_cells(mDec_IDEAL_cell, decisionWS_read)
    cell_values = [x if x is not None else 0 for x in cell_values]
    mod.write_XLS_cells(mDec_IDEAL_cell, cell_values, decisionWS_write)
    if (cell_values[7] == 'Ja'):
        cell_values[7] = 1
    else:
        cell_values[7] = 0
    mDec_IDEAL[:, co] = np.array(cell_values)
    if PERIOD<IDEAL_RD:
        mDec_IDEAL[[2, 3]] = 0
    if PERIOD<MARKET_2:
        mDec_IDEAL[[0, 4, 7]] = 0
    if PERIOD<MARKET_3:
        mDec_IDEAL[[1, 5, 6]] = 0
    cell_values = mod.read_XLS_cells(mDec_GESAMT_cell, decisionWS_read)
    cell_values = [x if x is not None else 0 for x in cell_values]
    mod.write_XLS_cells(mDec_GESAMT_cell, cell_values, decisionWS_write)
    if (cell_values[3] == 'Ja'):
        cell_values[3] = 1
    else:
        cell_values[3] = 0
    mDec_GESAMT[:, co] = np.array(cell_values)
    if PERIOD<MARKET_1:
        mDec_GESAMT[2] = 0
    cell_values = mod.read_XLS_cells(pDec_SOLID_cell, decisionWS_read)
    cell_values = [x if x is not None else 0 for x in cell_values]
    mod.write_XLS_cells(pDec_SOLID_cell, cell_values, decisionWS_write)
    pDec_SOLID[:, co] = np.array(cell_values)
    cell_values = mod.read_XLS_cells(pDec_IDEAL_cell, decisionWS_read)
    cell_values = [x if x is not None else 0 for x in cell_values]
    mod.write_XLS_cells(pDec_IDEAL_cell, cell_values, decisionWS_write)
    pDec_IDEAL[:, co] = np.array(cell_values)
    if PERIOD<MARKET_2:
        pDec_IDEAL[[0, 1, 2]] = 0
    pAll_TA_neu[:, co] = mod.read_XLS_range(pAll_TA_neu_cell, decisionWS_read)
    pAll_TA_alt[:, co] = mod.read_XLS_range(pAll_TA_alt_cell, decisionWS_read)
    mod.write_XLS_range(pAll_TA_neu_cell, pAll_TA_neu[:, co], decisionWS_write)
    mod.write_XLS_range(pAll_TA_alt_cell, pAll_TA_neu[:, co], decisionWS_write)
    for col in range(NUM_CELLS):
        if (pAll_TA_neu[col, co] != pAll_TA_alt[col, co]):
            if (pAll_TA_alt[col, co] == None):
                pAll_TA_inv[col, co] = 1
            elif (pAll_TA_alt[col, co] == '--'):
                pass
            else:
                pAll_TA_sell[col, co] = 1
            if   (pAll_TA_neu[col, co] == "Typ 'S' neu"):
                pAll_TA_buy[col, co] = 1
            elif (pAll_TA_neu[col, co] == "Typ 'M' neu"):
                pAll_TA_buy[col, co] = 2
            elif (pAll_TA_neu[col, co] == "Typ 'L' neu"):
                pAll_TA_buy[col, co] = 3
    del(col)
    cell_values = mod.read_XLS_cells(pDec_HR_cell, decisionWS_read)
    cell_values = [x if x is not None else 0 for x in cell_values]
    mod.write_XLS_cells(pDec_HR_cell, cell_values, decisionWS_write)
    pDec_HR[:, co] = np.array(cell_values)
    pDec_HR[2, co] = pDec_HR[2, co]/100
    pDec_HR[2, co] = max(pDec_HR[2, co],
                         max(pDec_HR_h[2, co, PERIOD+OFFSET-1] - 0.01, 0))
    cell_values = mod.read_XLS_cells(fDec_FIN_cell, decisionWS_read)
    cell_values = [x if x is not None else 0 for x in cell_values]
    mod.write_XLS_cells(fDec_FIN_cell, cell_values, decisionWS_write)
    if (cell_values[4] == None):
        cell_values[4] = 0
    fDec_FIN[:, co] = np.array(cell_values)
    decisionWS_write['P2'].value   = PERIOD+1
    decisionWS_write['AD28'].value = decisionWS_read['L28'].value
    decisionWS_write['AD23'].value = szenario['Sondermarkt'][1, PERIOD+OFFSET+1]
    type_TA = [None,
            f"'S'({PERIOD:0>2d})", f"'M'({PERIOD:0>2d})", f"'L'({PERIOD:0>2d})"]
    for sl in range(NUM_CELLS):
        if pAll_TA_buy[sl, co] != 0:
            pAll_TA_neu[sl, co] = type_TA[pAll_TA_buy[sl, co]]
            decisionWS_write.cell(row=14+sl, column=10
                                  ).value = type_TA[pAll_TA_buy[sl, co]]
            decisionWS_write.cell(row=14+sl, column=11
                                  ).value = type_TA[pAll_TA_buy[sl, co]]
    del(sl)
    decisionWB_write.save(filename = company_dir[co] + decision_files_new[co])
    decisionWB_read.close()
    decisionWB_write.close()
    mDec_SOLID_h[:, co, PERIOD+OFFSET] = mDec_SOLID[:, co]
    mDec_IDEAL_h[:, co, PERIOD+OFFSET] = mDec_IDEAL[:, co]
    mDec_GESAMT_h[:, co, PERIOD+OFFSET] = mDec_GESAMT[:, co]
    pDec_SOLID_h[0:3, co, PERIOD+OFFSET] = pDec_SOLID[:, co]
    pDec_IDEAL_h[0:3, co, PERIOD+OFFSET] = pDec_IDEAL[:, co]
    pDec_HR_h[0:4, co, PERIOD+OFFSET] = pDec_HR[:, co]
    fDec_FIN_h[0:5, co, PERIOD+OFFSET] = fDec_FIN[:, co]
    if (PERIOD == 0):
        reportWB_read  = load_workbook(
            filename = scenario_dir + report_files_old[co])
    else:
        reportWB_read  = load_workbook(
            filename = company_dir[co] + report_files_old[co])
    reportWB_read.save(filename = company_dir[co] + report_files[co])
    reportWB_read.close()
    if (PERIOD == 0):
        reportWB_read  = load_workbook(
            filename = scenario_dir + report_files_old[co],
            data_only=True)
    else:
        reportWB_read  = load_workbook(
            filename = company_dir[co] + report_files_old[co],
            data_only=True)
    reportWB_write = load_workbook(
        filename = company_dir[co] + report_files[co])
    reportWS_read  = reportWB_read['GuV']
    reportWS_write = reportWB_write['GuV']
    reportWS_write['C36'].value  = reportWS_read['C39'].value
    reportWS_read  = reportWB_read['Bilanz']
    reportWS_write = reportWB_write['Bilanz']
    read_range  = (('D5' , 'D8' ),
                   ('D11', 'D15'),
                   ('I5',  'I8' ),
                   ('I11', 'I15'))
    write_range = (('E5' , 'E8' ),
                   ('E11', 'E15'),
                   ('J5',  'J8' ),
                   ('J11', 'J15'))
    for ndx in range(len(read_range)):
        cell_values = mod.read_XLS_range(read_range[ndx], reportWS_read)
        mod.write_XLS_range(write_range[ndx], cell_values, reportWS_write)
    reportWS_read  = reportWB_read['Personal']
    reportWS_write = reportWB_write['Personal']
    read_range  = ('D9', 'F9')
    write_range = ('D5', 'F5')
    cell_values = mod.read_XLS_range(read_range, reportWS_read)
    mod.write_XLS_range(write_range, cell_values, reportWS_write)
    reportWS_write['F29'].value  = reportWS_read['F33'].value
    reportWS_read  = reportWB_read['Anlageverm�gen']
    reportWS_write = reportWB_write['Anlageverm�gen']
    read_range  = (( 'D15',  'D19'),
                   ( 'E15',  'E19'),
                   ( 'F15',  'F19'),
                   ( 'G15',  'G19'),
                   ( 'N11',  'U11'),
                   ( 'X15',  'X24'),
                   ('AF15', 'AF19'))
    write_range = (( 'D16',  'D20'),
                   ( 'E16',  'E20'),
                   ( 'F16',  'F20'),
                   ( 'G16',  'G20'),
                   ( 'N7' ,  'U7' ),
                   ( 'X16',  'X25'),
                   ('AF16', 'AF20'))
    for ndx in range(len(read_range)):
        cell_values = mod.read_XLS_range(read_range[ndx], reportWS_read)
        mod.write_XLS_range(write_range[ndx], cell_values, reportWS_write)
    reportWS_read  = reportWB_read['Vorr�te']
    reportWS_write = reportWB_write['Vorr�te']
    read_range  = (('C10', 'K10'),
                   ('C18', 'N18'))
    write_range = (('C6',  'K6'),
                   ('C14', 'N14'))
    for ndx in range(len(read_range)):
        cell_values = mod.read_XLS_range(read_range[ndx], reportWS_read)
        mod.write_XLS_range(write_range[ndx], cell_values, reportWS_write)
    reportWS_read  = reportWB_read['Finanzwerte']
    reportWS_write = reportWB_write['Finanzwerte']
    read_range  = (('C17', 'D17'),
                   ('K17', 'L17'),
                   ('O17', 'O21'),
                   ('P17', 'P21'),
                   ('S17', 'S26'),
                   ('T17', 'T26'))
    write_range = (('C18', 'D18'),
                   ('K18', 'L18'),
                   ('O18', 'O22'),
                   ('P18', 'P22'),
                   ('S18', 'S27'),
                   ('T18', 'T27'))
    for ndx in range(len(read_range)):
        cell_values = mod.read_XLS_range(read_range[ndx], reportWS_read)
        mod.write_XLS_range(write_range[ndx], cell_values, reportWS_write)
    reportWS_write['D7'].value  = reportWS_read['C7'].value
    reportWS_write['H7'].value  = reportWS_read['G7'].value
    reportWS_write['H18'].value  = reportWS_read['H17'].value
    reportWS_read  = reportWB_read['Aktuelle Wirtschaftsdaten']
    reportWS_write = reportWB_write['Aktuelle Wirtschaftsdaten']
    szenario_tuple = ('PreisHStoff', 'PreisMat1', 'PreisMat2', 'LagerMat',
                      'Zinsen', 'Finanzen',
                      'Gehalt', 'LohnNK', 'BedarfSOLID', 'BedarfIDEAL',
                      'AV_GG',
                      'AnlagenS', 'AnlagenM', 'AnlagenL',
                      'Sondermarkt', 'TransportFE', 'LagerFE')
    write_range = (('D6' , 'D10'),
                   ('D12', 'D17'),
                   ('D19', 'D24'),
                   ('D26', 'D28'),
                   ('D33', 'D38'),
                   ('D40', 'D42'),
                   ('I6',  'I8' ),
                   ('I10', 'I18'),
                   ('I23', 'I26'),
                   ('I28', 'I31'),
                   ('I36', 'I38'),
                   ('N6',  'N11'),
                   ('N13', 'N18'),
                   ('N20', 'N25'),
                   ('N29', 'N30'),
                   ('N32', 'N41'),
                   ('N43', 'N46'))
    for ndx in range(len(write_range)):
        mod.write_XLS_range(
            write_range[ndx],
            szenario[szenario_tuple[ndx]][:, PERIOD+OFFSET+1],
            reportWS_write)
    write_range = (('E6' , 'E10'),
                   ('E12', 'E17'),
                   ('E19', 'E24'),
                   ('E26', 'E28'),
                   ('E33', 'E38'),
                   ('E40', 'E42'),
                   ('J6',  'J8' ),
                   ('J10', 'J18'),
                   ('J23', 'J26'),
                   ('J28', 'J31'),
                   ('J36', 'J38'),
                   ('O6',  'O11'),
                   ('O13', 'O18'),
                   ('O20', 'O25'),
                   ('O29', 'O30'),
                   ('O32', 'O41'),
                   ('O43', 'O46'))
    for ndx in range(len(write_range)):
        mod.write_XLS_range(
            write_range[ndx],
            szenario[szenario_tuple[ndx]][:, PERIOD+OFFSET],
            reportWS_write)
    reportWS_write['I41'].value  = szenario['AV_BGA'][PERIOD+OFFSET+1]
    reportWS_write['J41'].value  = szenario['AV_BGA'][PERIOD+OFFSET]
    reportWB_write.save(filename = company_dir[co] + report_files[co])
    reportWS_write = reportWB_write['GuV']
    reportWS_write['F32'].value  = fDec_FIN[4, co]
    reportWS_write = reportWB_write['Fertigung']
    reportWS_write['D5'].value  = pDec_SOLID[0, co]
    reportWS_write['E5'].value  = pDec_IDEAL[0, co]
    reportWS_write = reportWB_write['Erl�sRechnung']
    reportWS_write['F10'].value = mDec_SOLID[0, co]
    reportWS_write['G11'].value = mDec_SOLID[1, co]
    reportWS_write['F23'].value = mDec_IDEAL[0, co]
    reportWS_write['G24'].value = mDec_IDEAL[1, co]
    reportWS_write = reportWB_write['Personal']
    reportWS_write['D10'].value = pDec_HR[0, co]
    reportWS_write['E9'].value  = (mDec_GESAMT[1, co]
                                    + mDec_GESAMT[2, co])
    reportWS_write['H14'].value = pDec_HR[2, co]
    reportWS_write['G23'].value = pDec_HR[1, co]
    reportWS_write = reportWB_write['KostenRechnung 1']
    reportWS_write['C52'].value  = szenario['AV_GG'][1, PERIOD+OFFSET]
    reportWS_write['C77'].value  = mDec_GESAMT[0, co]
    reportWS_write['C78'].value  = mDec_SOLID[4, co]
    reportWS_write['C79'].value  = mDec_SOLID[5, co]
    reportWS_write['C80'].value  = mDec_IDEAL[4, co]
    reportWS_write['C81'].value  = mDec_IDEAL[5, co]
    reportWS_write['C85'].value  = mDec_GESAMT[3, co]*COST_INDUSTRY_REPORT
    reportWS_write['C86'].value  = mDec_SOLID[7, co]*COST_MARKET_REPORT
    reportWS_write['C87'].value  = mDec_IDEAL[7, co]*COST_MARKET_REPORT
    reportWS_write = reportWB_write['Anlageverm�gen']
    reportWS_write['D15'].value  = mDec_SOLID[2, co]
    reportWS_write['E15'].value  = mDec_SOLID[3, co]
    reportWS_write['F15'].value  = mDec_IDEAL[2, co]
    reportWS_write['G15'].value  = mDec_IDEAL[3, co]
    reportWS_write['X15'].value  = (szenario['AV_BGA'][PERIOD+OFFSET]
                                    + pDec_HR[3, co])
    reportWS_write['AF15'].value = pDec_HR[1, co]
    type_TA = [None, "'S'", "'M'", "'L'"]
    ta_data = np.vstack((
        np.zeros(szenario['AnlagenS'][:, PERIOD+OFFSET].size),
        szenario['AnlagenS'][:, PERIOD+OFFSET],
        szenario['AnlagenM'][:, PERIOD+OFFSET],
        szenario['AnlagenL'][:, PERIOD+OFFSET]))
    n_cells = 0
    inv_TA = 0
    res_val = 0
    imp_loss = 0
    num_TA = 0
    age_TA = 0
    ke_TA = 0
    for sl in range(NUM_CELLS):
        reportWS_write.cell(row=8 , column=14+sl).value = 0
        reportWS_write.cell(row=9 , column=14+sl).value = 0
        if pAll_TA_inv[sl, co] == 1:
            reportWS_write.cell(row=13, column=14+sl).value = 1
            n_cells += 1
        if pAll_TA_sell[sl, co] == 1:
            res_val +=  (reportWS_write.cell(row=7 , column=14+sl).value
                          * reportWS_write.cell(row=23, column=14+sl).value)
            imp_loss += (reportWS_write.cell(row=7 , column=14+sl).value
                         * (1-reportWS_write.cell(row=23, column=14+sl).value))
            reportWS_write.cell(row=9 , column=14+sl
                ).value = -reportWS_write.cell(row=7 , column=14+sl).value
            reportWS_write.cell(row=15, column=14+sl).value = None
            reportWS_write.cell(row=16, column=14+sl).value = None
            reportWS_write.cell(row=17, column=14+sl).value = None
            reportWS_write.cell(row=18, column=14+sl).value = None
            reportWS_write.cell(row=19, column=14+sl).value = None
            reportWS_write.cell(row=21, column=14+sl).value = None
            reportWS_write.cell(row=22, column=14+sl).value = None
            reportWS_write.cell(row=23, column=14+sl).value = None
        if pAll_TA_buy[sl, co] != 0:
            reportWS_write.cell(row=8 , column=14+sl
                                ).value = ta_data[pAll_TA_buy[sl, co], 0]
            reportWS_write.cell(row=15, column=14+sl
                                ).value = type_TA[pAll_TA_buy[sl, co]]
            reportWS_write.cell(row=16, column=14+sl
                                ).value = PERIOD
            reportWS_write.cell(row=17, column=14+sl
                                ).value = ta_data[pAll_TA_buy[sl, co], 0]
            inv_TA += ta_data[pAll_TA_buy[sl, co], 0]
            reportWS_write.cell(row=18, column=14+sl
                                ).value = ta_data[pAll_TA_buy[sl, co], 1]
            reportWS_write.cell(row=19, column=14+sl
                                ).value = ta_data[pAll_TA_buy[sl, co], 2]
            reportWS_write.cell(row=21, column=14+sl
                                ).value = ta_data[pAll_TA_buy[sl, co], 3]
            reportWS_write.cell(row=22, column=14+sl
                                ).value = ta_data[pAll_TA_buy[sl, co], 4]
            reportWS_write.cell(row=23, column=14+sl
                                ).value = ta_data[pAll_TA_buy[sl, co], 5]
        if reportWS_write.cell(row=13, column=14+sl).value == 1:
            if reportWS_write.cell(row=16, column=14+sl).value != None:
                num_TA += 1
                age_TA += (PERIOD+1 - reportWS_write.cell(row=16, column=14+sl
                                                          ).value)
                ke_TA += reportWS_write.cell(row=22, column=14+sl).value
    pRes_TA_h[0, co, PERIOD+OFFSET] = ke_TA
    pRes_TA_h[2, co, PERIOD+OFFSET] = (inv_TA
                                       - res_val
                                       + n_cells
                                       * szenario['AV_GG'][2, PERIOD+OFFSET])
    del(sl, ta_data, ke_TA)
    reportWS_write['J8'].value = (szenario['AV_GG'][0, PERIOD+OFFSET]
                                  + n_cells*szenario['AV_GG'][2, PERIOD+OFFSET])
    reportWS_write['M25'].value = res_val
    reportWS_write['M26'].value = imp_loss
    reportWS_write = reportWB_write['Aktuelle Wirtschaftsdaten']
    reportWS_write['O2'].value   = PERIOD
    reportWS_write['I39'].value  = (reportWS_write['I39'].value
                                    + szenario['AV_GG'][0, PERIOD+OFFSET+1]/50
                                    + n_cells*szenario['AV_GG'][2, PERIOD+OFFSET]
                                    /UL_CELLS)
    reportWS_write['J39'].value  = (reportWS_write['J39'].value
                                    + n_cells*szenario['AV_GG'][2, PERIOD+OFFSET]
                                    /UL_CELLS)
    reportWS_write = reportWB_write['Finanzwerte']
    reportWS_write['C17'].value = fDec_FIN[3, co]
    reportWS_write['K17'].value = fDec_FIN[0, co]
    reportWS_write['O17'].value = fDec_FIN[1, co]
    reportWS_write['S17'].value = fDec_FIN[2, co]
    reportWS_write = reportWB_write['Vorr�te']
    reportWS_write['C7'].value  = (pDec_SOLID[1, co] + pDec_IDEAL[1, co])
    ndx = min(np.searchsorted(szenario['MengeHStoff'],
                              pDec_SOLID[1, co] + pDec_IDEAL[1, co]),
              len(szenario['PreisHStoff'][:, PERIOD+OFFSET])-1)
    reportWS_write['D7'].value = szenario['PreisHStoff'][ndx, PERIOD+OFFSET]
    reportWS_write['F7'].value  = pDec_SOLID[2, co]
    ndx = min(np.searchsorted(szenario['MengeMat1'], pDec_SOLID[2, co]),
              len(szenario['PreisMat1'][:, PERIOD+OFFSET])-1)
    reportWS_write['G7'].value = szenario['PreisMat1'][ndx, PERIOD+OFFSET]
    reportWS_write['I7'].value  = pDec_IDEAL[2, co]
    ndx = min(np.searchsorted(szenario['MengeMat2'], pDec_IDEAL[2, co]),
              len(szenario['PreisMat2'][:, PERIOD+OFFSET])-1)
    reportWS_write['J7'].value = szenario['PreisMat2'][ndx, PERIOD+OFFSET]
    reportWB_write.save(filename = company_dir[co] + report_files[co])
    reportWB_read.close()
    reportWB_write.close()
    prod_qual = mod.product_quality(
        np.vstack((mDec_SOLID_h[2:4, co, PERIOD+OFFSET-1:PERIOD+OFFSET+1]/1000,
                    mDec_IDEAL_h[2:4, co, PERIOD+OFFSET-1:PERIOD+OFFSET+1]/1000)),
        np.vstack((mMix_SOLID_h[0:4, co, PERIOD+OFFSET-2:PERIOD+OFFSET],
                    mMix_IDEAL_h[0:4, co, PERIOD+OFFSET-2:PERIOD+OFFSET])))
    mMix_SOLID_h[0:7, co, PERIOD+OFFSET] = prod_qual[0]
    mMix_IDEAL_h[0:7, co, PERIOD+OFFSET] = prod_qual[1]
    del(prod_qual)
    if PERIOD == MARKET_1:
        mMix_SOLID_h[ 8, co, PERIOD+OFFSET-1] = mDec_SOLID_h[5, co, PERIOD+OFFSET]
        mMix_SOLID_h[10, co, PERIOD+OFFSET-1] = mDec_GESAMT_h[2, co, PERIOD+OFFSET]
    if PERIOD == MARKET_2:
        mMix_IDEAL_h[ 7, co, PERIOD+OFFSET-1] = mDec_IDEAL_h[4, co, PERIOD+OFFSET]
        mMix_IDEAL_h[ 9, co, PERIOD+OFFSET-1] = mDec_GESAMT_h[1, co, PERIOD+OFFSET]
    if PERIOD == MARKET_3:
        mMix_IDEAL_h[ 8, co, PERIOD+OFFSET-1] = mDec_IDEAL_h[5, co, PERIOD+OFFSET]
        mMix_IDEAL_h[10, co, PERIOD+OFFSET-1] = mDec_GESAMT_h[2, co, PERIOD+OFFSET]
    mMix_SOLID_h[7:11, co, PERIOD+OFFSET] = mod.mMix_effects(
        np.vstack((mDec_SOLID_h[4:6, co, PERIOD+OFFSET-1:PERIOD+OFFSET+1],
                   mDec_GESAMT_h[1:3, co, PERIOD+OFFSET-1:PERIOD+OFFSET+1])),
        mMix_SOLID_h[7:11, co, PERIOD+OFFSET-1])
    if PERIOD >= MARKET_2:
        mMix_IDEAL_h[7:11, co, PERIOD+OFFSET] = mod.mMix_effects(
            np.vstack((mDec_IDEAL_h[4:6, co, PERIOD+OFFSET-1:PERIOD+OFFSET+1],
                       mDec_GESAMT_h[1:3, co, PERIOD+OFFSET-1:PERIOD+OFFSET+1])),
            mMix_IDEAL_h[7:11, co, PERIOD+OFFSET-1])
    cSat_SOLID_h[[0, 3, 4, 5, 8, 9], co, PERIOD+OFFSET] = mod.cSAT_SOLID(
        mDec_SOLID_h[0:2, co, PERIOD+OFFSET-1:PERIOD+OFFSET+1],
        [mMix_SOLID_h[6, co, PERIOD+OFFSET],
          szenario['eQI_SOLID'][PERIOD+OFFSET]],
        cSat_SOLID_h[:, co, PERIOD+OFFSET-2:PERIOD+OFFSET],
        mMix_SOLID_h[12:14, co, PERIOD+OFFSET-1])
    cSat_IDEAL_h[[0, 3, 4, 5, 8, 9], co, PERIOD+OFFSET] = mod.cSAT_IDEAL(
        mDec_IDEAL_h[0:2, co, PERIOD+OFFSET-1:PERIOD+OFFSET+1],
        [mMix_IDEAL_h[6, co, PERIOD+OFFSET],
          szenario['eQI_IDEAL'][PERIOD+OFFSET]],
        cSat_IDEAL_h[:, co, PERIOD+OFFSET-2:PERIOD+OFFSET],
        mMix_IDEAL_h[12:14, co, PERIOD+OFFSET-1])
    mMix_SOLID_h[11:14, co, PERIOD+OFFSET] = mod.brand_strength(
        mDec_GESAMT_h[0, co, PERIOD+OFFSET-2:PERIOD+OFFSET+1],
        cSat_SOLID_h[[4, 9], co, PERIOD+OFFSET],
        mMix_SOLID_h[11, co, PERIOD+OFFSET-1])
    mMix_IDEAL_h[11:14, co, PERIOD+OFFSET] = mod.brand_strength(
        mDec_GESAMT_h[0, co, PERIOD+OFFSET-2:PERIOD+OFFSET+1],
        cSat_IDEAL_h[[4, 9], co, PERIOD+OFFSET],
        mMix_IDEAL_h[11, co, PERIOD+OFFSET-1])
    (pDec_eff, pHR_ar0) = mod.pDec_effects(
        pDec_HR_h[:, co, PERIOD+OFFSET-1:PERIOD+OFFSET+1],
        pRes_HR_h[0:12, co, PERIOD+OFFSET-2:PERIOD+OFFSET+1])
    pRes_HR_h[8:12, co, PERIOD+OFFSET] = pDec_eff
    pRes_HR_h[[0, 3, 4, 6], co, PERIOD+OFFSET] = pHR_ar0
    pRes_HR_h[16, co, PERIOD+OFFSET] = mod.bs_emp(
        np.vstack((
            pRes_HR_h[8:12, co, PERIOD+OFFSET-1:PERIOD+OFFSET+1],
            szenario['eMA'][:, PERIOD+OFFSET-1:PERIOD+OFFSET+1],
            fRes_COMP_h[[0, 4, 7, 8], co, PERIOD+OFFSET-2:PERIOD+OFFSET],
            pRes_HR_h[16, co, PERIOD+OFFSET-1:PERIOD+OFFSET+1])))
    pRes_HR_h[15, co, PERIOD+OFFSET] = mod.pe_prod(
        pRes_HR_h[10, co, PERIOD+OFFSET])
    pRes_HR_h[7, co, PERIOD+OFFSET] = round(age_TA/num_TA, 1)
    info_period_WS  = info_period_WB['Branchenbericht']
    info_period_WS.cell(row=48, column=4+co
                        ).value = pDec_HR[2, co]
    info_period_WS.cell(row=49, column=4+co
                        ).value = pDec_HR[1, co]
    info_period_WS.cell(row=56, column=4+co
                        ).value = pRes_TA_h[0, co, PERIOD+OFFSET]
    info_period_WS.cell(row=58, column=4+co
                        ).value = pRes_TA_h[2, co, PERIOD+OFFSET]
    info_period_WS.cell(row=59, column=4+co
                        ).value = pDec_HR[3, co]
    info_period_WS.cell(row=60, column=4+co
                        ).value = (  mDec_SOLID[2, co]
                                   + mDec_SOLID[3, co]
                                   + mDec_IDEAL[2, co]
                                   + mDec_IDEAL[3, co])
    info_period_WS.cell(row=63, column=4+co
                        ).value = pRes_HR_h[16, co, PERIOD+OFFSET]-100
    info_period_WS  = info_period_WB['Marktbericht SOLID']
    info_period_WS.cell(row= 5, column=4+co
                        ).value = mMix_SOLID_h[4, co, PERIOD+OFFSET]-100
    info_period_WS.cell(row= 6, column=4+co
                        ).value = mMix_SOLID_h[5, co, PERIOD+OFFSET]-100
    info_period_WS.cell(row= 9, column=4+co
                        ).value = mDec_SOLID[0, co]
    info_period_WS.cell(row=10, column=4+co
                        ).value = mDec_SOLID[4, co]
    info_period_WS.cell(row=11, column=4+co
                        ).value = mDec_GESAMT[1, co]
    info_period_WS.cell(row=12, column=4+co
                        ).value = cSat_SOLID_h[4, co, PERIOD+OFFSET]-100
    info_period_WS.cell(row=13, column=4+co
                        ).value = mMix_SOLID_h[12, co, PERIOD+OFFSET]-100
    info_period_WS.cell(row=31, column=4+co
                        ).value = mDec_GESAMT[0, co]
    info_period_WS.cell(row=32, column=4+co
                        ).value = mDec_SOLID[8, co]
    info_period_WS.cell(row=33, column=4+co
                        ).value = szenario['Sondermarkt'][0, PERIOD+OFFSET]
    if PERIOD>=MARKET_1:
        info_period_WS.cell(row=20, column=4+co
                            ).value = mDec_SOLID[1, co]
        info_period_WS.cell(row=21, column=4+co
                            ).value = mDec_SOLID[5, co]
        info_period_WS.cell(row=22, column=4+co
                            ).value = mDec_GESAMT[2, co]
        info_period_WS.cell(row=23, column=4+co
                            ).value = cSat_SOLID_h[9, co, PERIOD+OFFSET]-100
        info_period_WS.cell(row=24, column=4+co
                            ).value = mMix_SOLID_h[13, co, PERIOD+OFFSET]-100
    if PERIOD>=MARKET_2:
        info_period_WS  = info_period_WB['Marktbericht IDEAL']
        info_period_WS.cell(row= 5, column=4+co
                            ).value = mMix_IDEAL_h[4, co, PERIOD+OFFSET]-100
        info_period_WS.cell(row= 6, column=4+co
                            ).value = mMix_IDEAL_h[5, co, PERIOD+OFFSET]-100
        info_period_WS.cell(row= 9, column=4+co
                            ).value = mDec_IDEAL[0, co]
        info_period_WS.cell(row=10, column=4+co
                            ).value = mDec_IDEAL[4, co]
        info_period_WS.cell(row=11, column=4+co
                            ).value = mDec_GESAMT[1, co]
        info_period_WS.cell(row=12, column=4+co
                            ).value = cSat_IDEAL_h[4, co, PERIOD+OFFSET]-100
        info_period_WS.cell(row=13, column=4+co
                            ).value = mMix_IDEAL_h[12, co, PERIOD+OFFSET]-100
        info_period_WS.cell(row=31, column=4+co
                            ).value = mDec_GESAMT[0, co]
    if PERIOD>=MARKET_3:
        info_period_WS.cell(row=20, column=4+co
                            ).value = mDec_IDEAL[1, co]
        info_period_WS.cell(row=21, column=4+co
                            ).value = mDec_IDEAL[5, co]
        info_period_WS.cell(row=22, column=4+co
                            ).value = mDec_GESAMT[2, co]
        info_period_WS.cell(row=23, column=4+co
                            ).value = cSat_IDEAL_h[9, co, PERIOD+OFFSET]-100
        info_period_WS.cell(row=24, column=4+co
                            ).value = mMix_IDEAL_h[13, co, PERIOD+OFFSET]-100
if (PERIOD == 0):
    for co in range(1, num_companies0):
        mDec_SOLID_h[:, co,
                     PERIOD+OFFSET] = mDec_SOLID_h[:, 0, PERIOD+OFFSET]
        mDec_IDEAL_h[:, co,
                     PERIOD+OFFSET] = mDec_IDEAL_h[:, 0, PERIOD+OFFSET]
        mDec_GESAMT_h[:, co,
                      PERIOD+OFFSET] = mDec_GESAMT_h[:, 0, PERIOD+OFFSET]
        mMix_SOLID_h[:, co,
                     PERIOD+OFFSET] = mMix_SOLID_h[:, 0, PERIOD+OFFSET]
        cSat_SOLID_h[:, co,
                     PERIOD+OFFSET] = cSat_SOLID_h[:, 0, PERIOD+OFFSET]
        mMix_IDEAL_h[:, co,
                     PERIOD+OFFSET] = mMix_IDEAL_h[:, 0, PERIOD+OFFSET]
        cSat_IDEAL_h[:, co,
                     PERIOD+OFFSET] = cSat_IDEAL_h[:, 0, PERIOD+OFFSET]
        pDec_SOLID_h[:, co,
                     PERIOD+OFFSET] = pDec_SOLID_h[:, 0, PERIOD+OFFSET]
        pDec_IDEAL_h[:, co,
                     PERIOD+OFFSET] = pDec_IDEAL_h[:, 0, PERIOD+OFFSET]
        pRes_TA_h[:, co,
                  PERIOD+OFFSET] = pRes_TA_h[:, 0, PERIOD+OFFSET]
        pDec_HR_h[:, co,
                  PERIOD+OFFSET] = pDec_HR_h[:, 0, PERIOD+OFFSET]
        pRes_HR_h[8:12, co,
                  PERIOD+OFFSET] = pRes_HR_h[8:12, 0, PERIOD+OFFSET]
        pRes_HR_h[[0, 3, 4, 6], co,
                  PERIOD+OFFSET] = pRes_HR_h[[0, 3, 4, 6], 0, PERIOD+OFFSET]
        pRes_HR_h[16, co, PERIOD+OFFSET] = pRes_HR_h[16, 0, PERIOD+OFFSET]
        pRes_HR_h[15, co, PERIOD+OFFSET] = pRes_HR_h[15, 0, PERIOD+OFFSET]
        pRes_HR_h[ 7, co, PERIOD+OFFSET] = pRes_HR_h[ 7, 0, PERIOD+OFFSET]
        fDec_FIN_h[:, co,
                   PERIOD+OFFSET] = fDec_FIN_h[:, 0, PERIOD+OFFSET]
        fDec_FIN_h[:, co,
                   PERIOD+OFFSET] = fDec_FIN_h[:, 0, PERIOD+OFFSET]
        fRes_COMP_h[:, co,
                    PERIOD+OFFSET] = fRes_COMP_h[:, 0, PERIOD+OFFSET]
        info_period_WS  = info_period_WB['Branchenbericht']
        info_period_WS.cell(row=48, column=4+co
                            ).value = pDec_HR[2, 0]
        info_period_WS.cell(row=49, column=4+co
                            ).value = pDec_HR[1, 0]
        info_period_WS.cell(row=56, column=4+co
                            ).value = pRes_TA_h[0, co, PERIOD+OFFSET]
        info_period_WS.cell(row=58, column=4+co
                            ).value = pRes_TA_h[2, co, PERIOD+OFFSET]
        info_period_WS.cell(row=59, column=4+co
                            ).value = pDec_HR[3, 0]
        info_period_WS.cell(row=60, column=4+co
                            ).value = (  mDec_SOLID[2, 0]
                                       + mDec_SOLID[3, 0]
                                       + mDec_IDEAL[2, 0]
                                       + mDec_IDEAL[3, 0])
        info_period_WS.cell(row=63, column=4+co
                            ).value = pRes_HR_h[16, 0, PERIOD+OFFSET]-100
        info_period_WS  = info_period_WB['Marktbericht SOLID']
        info_period_WS.cell(row= 5, column=4+co
                            ).value = mMix_SOLID_h[4, 0, PERIOD+OFFSET]-100
        info_period_WS.cell(row= 6, column=4+co
                            ).value = mMix_SOLID_h[5, 0, PERIOD+OFFSET]-100
        info_period_WS.cell(row= 9, column=4+co
                            ).value = mDec_SOLID[0, 0]
        info_period_WS.cell(row=10, column=4+co
                            ).value = mDec_SOLID[4, 0]
        info_period_WS.cell(row=11, column=4+co
                            ).value = mDec_GESAMT[1, 0]
        info_period_WS.cell(row=12, column=4+co
                            ).value = cSat_SOLID_h[4, 0, PERIOD+OFFSET]-100
        info_period_WS.cell(row=13, column=4+co
                            ).value = mMix_SOLID_h[12, 0, PERIOD+OFFSET]-100
        info_period_WS.cell(row=31, column=4+co
                            ).value = mDec_GESAMT[0, 0]
        info_period_WS.cell(row=32, column=4+co
                            ).value = mDec_SOLID[8, 0]
        info_period_WS.cell(row=33, column=4+co
                            ).value = szenario['Sondermarkt'][0, PERIOD+OFFSET]
    del(co)
del(COST_INDUSTRY_REPORT, COST_MARKET_REPORT)
del(mDec_SOLID_cell, mDec_IDEAL_cell, mDec_GESAMT_cell,
    pDec_SOLID_cell, pDec_IDEAL_cell, pDec_HR_cell,
    pAll_TA_neu_cell, pAll_TA_alt_cell, fDec_FIN_cell)
del(pAll_TA_neu, pAll_TA_alt)
del(cell_values)
del(read_range, write_range, szenario_tuple)
del(decision_dir, decision_template)
del(decisionWB_read, decisionWB_write,
    decisionWS_read, decisionWS_write)
del(reportWS_read, reportWB_read, report_files_old)
del(n_cells, inv_TA, res_val, imp_loss, num_TA, age_TA)
del(pDec_eff, pHR_ar0)
mean_time = time.time()
print('')
print('1. Schleife')
print(round(mean_time - start_time, 1), f'({mean_time - mean_time0:.1f})')
mean_time0 = mean_time
pRes_HR_h[17, :, PERIOD+OFFSET] = mod.emp_motivation(
    np.vstack((
        pRes_HR_h[8:12, :, PERIOD+OFFSET],
        pRes_HR_h[12, :, PERIOD+OFFSET-1],
        pRes_HR_h[16, :, PERIOD+OFFSET],
        pRes_HR_h[17, :, PERIOD+OFFSET-1])))
pRes_HR_h[18:21, :, PERIOD+OFFSET] = mod.em_effects(
    pRes_HR_h[17, :, PERIOD+OFFSET])
pRes_HR_h[14, :, PERIOD+OFFSET] = mod.lm_prod(
    np.vstack((
        pRes_HR_h[8:12, :, PERIOD+OFFSET],
        pRes_HR_h[14, :, PERIOD+OFFSET-1])))
lm_supply = mod.lm_supply(
    pRes_HR_h[8, :, PERIOD+OFFSET],
    [szenario['eMA'][0, PERIOD+OFFSET],
     szenario['LohnNK'][5, PERIOD+OFFSET]])
pRes_HR_h[13, :, PERIOD+OFFSET] = lm_supply
lm_share = mod.lm_share(
    pRes_HR_h[8:12, :, PERIOD+OFFSET])
(hr_dep, emp_prod) = mod.hr_department(
    np.vstack((pRes_HR_h[1, :, PERIOD+OFFSET-1],
               pDec_HR_h[0, :, PERIOD+OFFSET])),
    np.vstack((pRes_HR_h[13:21, :, PERIOD+OFFSET],
               pRes_HR_h[21, :, PERIOD+OFFSET-1],
               lm_share)),
    szenario['LohnNK'][7, PERIOD+OFFSET])
pRes_HR_h[ 1, :, PERIOD+OFFSET] = hr_dep[0]
pRes_HR_h[ 2, :, PERIOD+OFFSET] = hr_dep[1]
pRes_HR_h[21, :, PERIOD+OFFSET] = emp_prod
del(hr_dep, emp_prod, lm_share, lm_supply)
(mv_cur, arma0) = mod.sm_volume(
    np.hstack((np.array([mMix_SOLID_h[6, :, PERIOD+OFFSET]]).T,
               mDec_SOLID_h[ 0, :, PERIOD+OFFSET-1:PERIOD+OFFSET+1],
               mMix_SOLID_h[ 7, :, PERIOD+OFFSET-1:PERIOD+OFFSET+1],
               mMix_SOLID_h[ 9, :, PERIOD+OFFSET-1:PERIOD+OFFSET+1],
               np.array([mMix_SOLID_h[12, :, PERIOD+OFFSET]]).T)),
    np.vstack((szenario['MarktVol'][[0, 4, 8, 12], PERIOD+OFFSET-1:PERIOD+OFFSET+1],
               szenario['Preisindex'][PERIOD+OFFSET-1:PERIOD+OFFSET+1],
               szenario['eQI_SOLID'][PERIOD+OFFSET-1:PERIOD+OFFSET+1])),
    0, szenario['MarktVol'][1:4, PERIOD+OFFSET-1])
szenario['MarktVol'][1:4, PERIOD+OFFSET] = arma0
prod_demand[0] = mod.sm_share(
    np.vstack((mMix_SOLID_h[ 6, :, PERIOD+OFFSET],
               mDec_SOLID_h[ 0, :, PERIOD+OFFSET],
               mMix_SOLID_h[ 7, :, PERIOD+OFFSET],
               mMix_SOLID_h[ 9, :, PERIOD+OFFSET],
               mMix_SOLID_h[12, :, PERIOD+OFFSET],
               cSat_SOLID_h[ 4, :, PERIOD+OFFSET])),
    mv_cur)
if PERIOD>=MARKET_1:
    (mv_cur, arma0) = mod.sm_volume(
        np.hstack((np.array([mMix_SOLID_h[6, :, PERIOD+OFFSET]]).T,
                    mDec_SOLID_h[ 1, :, PERIOD+OFFSET-1:PERIOD+OFFSET+1],
                    mMix_SOLID_h[ 8, :, PERIOD+OFFSET-1:PERIOD+OFFSET+1],
                    mMix_SOLID_h[10, :, PERIOD+OFFSET-1:PERIOD+OFFSET+1],
                    np.array([mMix_SOLID_h[13, :, PERIOD+OFFSET]]).T)),
        np.vstack((szenario['MarktVol'][[0, 4, 8, 12], PERIOD+OFFSET-1:PERIOD+OFFSET+1],
                    szenario['Preisindex'][PERIOD+OFFSET-1:PERIOD+OFFSET+1],
                    szenario['eQI_SOLID'][PERIOD+OFFSET-1:PERIOD+OFFSET+1])),
        1, szenario['MarktVol'][5:8, PERIOD+OFFSET-1])
    szenario['MarktVol'][5:8, PERIOD+OFFSET] = arma0
    prod_demand[1] = mod.sm_share(
        np.vstack((mMix_SOLID_h[ 6, :, PERIOD+OFFSET],
                   mDec_SOLID_h[ 1, :, PERIOD+OFFSET],
                   mMix_SOLID_h[ 8, :, PERIOD+OFFSET],
                   mMix_SOLID_h[10, :, PERIOD+OFFSET],
                   mMix_SOLID_h[13, :, PERIOD+OFFSET],
                   cSat_SOLID_h[ 9, :, PERIOD+OFFSET])),
        mv_cur/szenario['Finanzen'][0, PERIOD+OFFSET])
if PERIOD>=MARKET_2:
    (mv_cur, arma0) = mod.sm_volume(
        np.hstack((np.array([mMix_IDEAL_h[6, :, PERIOD+OFFSET]]).T,
                    mDec_IDEAL_h[ 0, :, PERIOD+OFFSET-1:PERIOD+OFFSET+1],
                    mMix_IDEAL_h[ 7, :, PERIOD+OFFSET-1:PERIOD+OFFSET+1],
                    mMix_IDEAL_h[ 9, :, PERIOD+OFFSET-1:PERIOD+OFFSET+1],
                    np.array([mMix_IDEAL_h[12, :, PERIOD+OFFSET]]).T)),
        np.vstack((szenario['MarktVol'][[0, 4, 8, 12], PERIOD+OFFSET-1:PERIOD+OFFSET+1],
                    szenario['Preisindex'][PERIOD+OFFSET-1:PERIOD+OFFSET+1],
                    szenario['eQI_IDEAL'][PERIOD+OFFSET-1:PERIOD+OFFSET+1])),
        2, szenario['MarktVol'][9:12, PERIOD+OFFSET-1])
    szenario['MarktVol'][9:12, PERIOD+OFFSET] = arma0
    prod_demand[2] = mod.sm_share(
        np.vstack((mMix_IDEAL_h[ 6, :, PERIOD+OFFSET],
                   mDec_IDEAL_h[ 0, :, PERIOD+OFFSET],
                   mMix_IDEAL_h[ 7, :, PERIOD+OFFSET],
                   mMix_IDEAL_h[ 9, :, PERIOD+OFFSET],
                   mMix_IDEAL_h[12, :, PERIOD+OFFSET],
                   cSat_IDEAL_h[ 4, :, PERIOD+OFFSET])),
        mv_cur)
if PERIOD>=MARKET_3:
    (mv_cur, arma0) = mod.sm_volume(
        np.hstack((np.array([mMix_IDEAL_h[6, :, PERIOD+OFFSET]]).T,
                    mDec_IDEAL_h[ 1, :, PERIOD+OFFSET-1:PERIOD+OFFSET+1],
                    mMix_IDEAL_h[ 8, :, PERIOD+OFFSET-1:PERIOD+OFFSET+1],
                    mMix_IDEAL_h[10, :, PERIOD+OFFSET-1:PERIOD+OFFSET+1],
                    np.array([mMix_IDEAL_h[13, :, PERIOD+OFFSET]]).T)),
        np.vstack((szenario['MarktVol'][[0, 4, 8, 12], PERIOD+OFFSET-1:PERIOD+OFFSET+1],
                    szenario['Preisindex'][PERIOD+OFFSET-1:PERIOD+OFFSET+1],
                    szenario['eQI_IDEAL'][PERIOD+OFFSET-1:PERIOD+OFFSET+1])),
        3, szenario['MarktVol'][13:16, PERIOD+OFFSET-1])
    szenario['MarktVol'][13:16, PERIOD+OFFSET] = arma0
    prod_demand[3] = mod.sm_share(
        np.vstack((mMix_IDEAL_h[ 6, :, PERIOD+OFFSET],
                   mDec_IDEAL_h[ 1, :, PERIOD+OFFSET],
                   mMix_IDEAL_h[ 8, :, PERIOD+OFFSET],
                   mMix_IDEAL_h[10, :, PERIOD+OFFSET],
                   mMix_IDEAL_h[13, :, PERIOD+OFFSET],
                   cSat_IDEAL_h[ 9, :, PERIOD+OFFSET])),
        mv_cur/szenario['Finanzen'][0, PERIOD+OFFSET])
del(mv_cur, arma0)
cSat_SOLID_h[ 2, :, PERIOD+OFFSET] = prod_demand[0]
cSat_SOLID_h[ 7, :, PERIOD+OFFSET] = prod_demand[1]
cSat_IDEAL_h[ 2, :, PERIOD+OFFSET] = prod_demand[2]
cSat_IDEAL_h[ 7, :, PERIOD+OFFSET] = prod_demand[3]
print('')
print('Nachfrage SOLID-Inland :', prod_demand[0])
print('Nachfrage SOLID-Ausland:', prod_demand[1])
print('Nachfrage IDEAL-Inland :', prod_demand[2])
print('Nachfrage IDEAL-Ausland:', prod_demand[3])
(prod_act, cap_ut) = mod.sm_supply(
    np.vstack((pDec_SOLID_h[ 0, :, PERIOD+OFFSET],
               pDec_IDEAL_h[ 0, :, PERIOD+OFFSET])),
    np.vstack((pRes_HR_h[ 1, :, PERIOD+OFFSET],
               pRes_HR_h[21, :, PERIOD+OFFSET],
               pRes_TA_h[ 0, :, PERIOD+OFFSET])),
    np.array([szenario['LohnNK'][ 1, PERIOD+OFFSET],
              szenario['BedarfSOLID'][ 2, PERIOD+OFFSET],
              szenario['BedarfIDEAL'][ 2, PERIOD+OFFSET],
              szenario['BedarfSOLID'][ 3, PERIOD+OFFSET],
              szenario['BedarfIDEAL'][ 3, PERIOD+OFFSET]]))
pDec_SOLID_h[ 3, :, PERIOD+OFFSET] = prod_act[0]
pDec_IDEAL_h[ 3, :, PERIOD+OFFSET] = prod_act[1]
pRes_HR_h[12, :, PERIOD+OFFSET] = cap_ut[0]
pRes_TA_h[ 1, :, PERIOD+OFFSET] = cap_ut[1]
del(cap_ut)
print('')
print('Fertigung SOLID:', prod_act[0])
print('Fertigung IDEAL:', prod_act[1])
prod_special = np.amin(
    np.vstack(((mMix_SOLID_h[14, :, PERIOD+OFFSET-1]
              + mMix_SOLID_h[15, :, PERIOD+OFFSET-1]
              + prod_act[0]),
                mDec_SOLID_h[ 8, :, PERIOD+OFFSET])), axis=0)
prod_special = np.amax(
    np.vstack((prod_special,
               np.zeros(num_companies0))), axis=0)
mDec_SOLID_h[ 8, :, PERIOD+OFFSET] = prod_special
del(prod_special)
prod_abroad = np.amin(
    np.vstack(((mMix_SOLID_h[14, :, PERIOD+OFFSET-1]
              + mMix_SOLID_h[15, :, PERIOD+OFFSET-1]
              + prod_act[0]
              - mDec_SOLID_h[ 8, :, PERIOD+OFFSET]),
                mDec_SOLID_h[ 6, :, PERIOD+OFFSET])), axis=0)
prod_abroad = np.amax(
    np.vstack((prod_abroad,
               np.zeros(num_companies0))), axis=0)
mDec_SOLID_h[ 6, :, PERIOD+OFFSET] = prod_abroad
prod_abroad = np.amin(
    np.vstack(((mMix_IDEAL_h[14, :, PERIOD+OFFSET-1]
                + mMix_IDEAL_h[15, :, PERIOD+OFFSET-1]
                + prod_act[1]),
                mDec_IDEAL_h[ 6, :, PERIOD+OFFSET])), axis=0)
prod_abroad = np.amax(
    np.vstack((prod_abroad,
               np.zeros(num_companies0))), axis=0)
mDec_IDEAL_h[ 6, :, PERIOD+OFFSET] = prod_abroad
del(prod_abroad)
prod_supply[0] = (mMix_SOLID_h[14, :, PERIOD+OFFSET-1]
                + prod_act[0]
                - mDec_SOLID_h[ 6, :, PERIOD+OFFSET]
                - mDec_SOLID_h[ 8, :, PERIOD+OFFSET])
prod_supply[1] = (mMix_SOLID_h[15, :, PERIOD+OFFSET-1]
                + mDec_SOLID_h[ 6, :, PERIOD+OFFSET])
prod_supply[2] = (mMix_IDEAL_h[14, :, PERIOD+OFFSET-1]
                + prod_act[1]
                - mDec_IDEAL_h[ 6, :, PERIOD+OFFSET])
prod_supply[3] = (mMix_IDEAL_h[15, :, PERIOD+OFFSET-1]
                + mDec_IDEAL_h[ 6, :, PERIOD+OFFSET])
prod_excess = prod_supply - prod_demand
mismatch_quant  = np.prod(prod_excess[0:2], axis=0)<0
mismatch_market = np.greater(prod_excess[0], np.zeros(num_companies0))
express_SOLID   = (np.amin(abs(prod_excess[0:2]), axis=0)
                   *mismatch_quant
                   *(2*mismatch_market-1))
mismatch_quant  = np.prod(prod_excess[2:4], axis=0)<0
mismatch_market = np.greater(prod_excess[2], np.zeros(num_companies0))
express_IDEAL   = (np.amin(abs(prod_excess[2:4]), axis=0)
                   *mismatch_quant
                   *(2*mismatch_market-1))
del(mismatch_quant, mismatch_market)
prod_supply[0] = prod_supply[0] - express_SOLID
prod_supply[1] = prod_supply[1] + express_SOLID
prod_supply[2] = prod_supply[2] - express_IDEAL
prod_supply[3] = prod_supply[3] + express_IDEAL
sales_cur[0] = np.amin(
    np.vstack((prod_supply[0],
               prod_demand[0])), axis = 0)
sales_cur[1] = np.amin(
    np.vstack((prod_supply[1],
               prod_demand[1])), axis = 0)
sales_cur[2] = np.amin(
    np.vstack((prod_supply[2],
               prod_demand[2])), axis = 0)
sales_cur[3] = np.amin(
    np.vstack((prod_supply[3],
               prod_demand[3])), axis = 0)
print('')
print('Absatz SOLID-Inland :', sales_cur[0])
print('Absatz SOLID-Ausland:', sales_cur[1])
print('Absatz IDEAL-Inland :', sales_cur[2])
print('Absatz IDEAL-Ausland:', sales_cur[3])
mMix_SOLID_h[14, :, PERIOD+OFFSET] = (mMix_SOLID_h[14, :, PERIOD+OFFSET-1]
                                    + prod_act[0]
                                    - mDec_SOLID_h[ 6, :, PERIOD+OFFSET]
                                    - express_SOLID
                                    - mDec_SOLID_h[ 8, :, PERIOD+OFFSET]
                                    - sales_cur[0])
mMix_SOLID_h[15, :, PERIOD+OFFSET] = (mMix_SOLID_h[15, :, PERIOD+OFFSET-1]
                                    + mDec_SOLID_h[ 6, :, PERIOD+OFFSET]
                                    + express_SOLID
                                    - sales_cur[1])
mMix_IDEAL_h[14, :, PERIOD+OFFSET] = (mMix_IDEAL_h[14, :, PERIOD+OFFSET-1]
                                    + prod_act[1]
                                    - mDec_IDEAL_h[ 6, :, PERIOD+OFFSET]
                                    - express_IDEAL
                                    - sales_cur[2])
mMix_IDEAL_h[15, :, PERIOD+OFFSET] = (mMix_IDEAL_h[15, :, PERIOD+OFFSET-1]
                                    + mDec_IDEAL_h[ 6, :, PERIOD+OFFSET]
                                    + express_IDEAL
                                    - sales_cur[3])
cSat_SOLID_h[ 1, :, PERIOD+OFFSET] = sales_cur[0]
cSat_SOLID_h[ 6, :, PERIOD+OFFSET] = sales_cur[1]
cSat_IDEAL_h[ 1, :, PERIOD+OFFSET] = sales_cur[2]
cSat_IDEAL_h[ 6, :, PERIOD+OFFSET] = sales_cur[3]
del(prod_act, prod_demand, prod_supply, prod_excess)
info_period_WS  = info_period_WB['Branchenbericht']
for co in range(num_companies0):
    info_period_WS.cell(row=50, column=4+co
                        ).value = pRes_HR_h[17, co, PERIOD+OFFSET]-100
    info_period_WS.cell(row=51, column=4+co
                        ).value = pRes_HR_h[21, co, PERIOD+OFFSET]-1
    info_period_WS.cell(row=52, column=4+co
                        ).value = pRes_HR_h[1 , co, PERIOD+OFFSET]
    info_period_WS.cell(row=53, column=4+co
                        ).value = pRes_HR_h[12, co, PERIOD+OFFSET]
    info_period_WS.cell(row=57, column=4+co
                        ).value = pRes_TA_h[1 , co, PERIOD+OFFSET]
info_period_WS  = info_period_WB['Marktbericht SOLID']
for co in range(num_companies0):
    info_period_WS.cell(row=14, column=4+co
                        ).value = cSat_SOLID_h[2, co, PERIOD+OFFSET]
    info_period_WS.cell(row=15, column=4+co
                        ).value = sales_cur[0, co]
    info_period_WS.cell(row=16, column=4+co
                        ).value = mMix_SOLID_h[14, co, PERIOD+OFFSET]
    if PERIOD>=MARKET_1:
        info_period_WS.cell(row=25, column=4+co
                            ).value = cSat_SOLID_h[7, co, PERIOD+OFFSET]
        info_period_WS.cell(row=26, column=4+co
                            ).value = sales_cur[1, co]
        info_period_WS.cell(row=27, column=4+co
                            ).value = mMix_SOLID_h[15, co, PERIOD+OFFSET]
if PERIOD>=MARKET_2:
    info_period_WS  = info_period_WB['Marktbericht IDEAL']
    for co in range(num_companies0):
        info_period_WS.cell(row=14, column=4+co
                            ).value = cSat_IDEAL_h[2, co, PERIOD+OFFSET]
        info_period_WS.cell(row=15, column=4+co
                            ).value = sales_cur[2, co]
        info_period_WS.cell(row=16, column=4+co
                            ).value = mMix_IDEAL_h[14, co, PERIOD+OFFSET]
        if PERIOD>=MARKET_3:
            info_period_WS.cell(row=25, column=4+co
                                ).value = cSat_IDEAL_h[7, co, PERIOD+OFFSET]
            info_period_WS.cell(row=26, column=4+co
                                ).value = sales_cur[3, co]
            info_period_WS.cell(row=27, column=4+co
                                ).value = mMix_IDEAL_h[15, co, PERIOD+OFFSET]
mean_time = time.time()
print('')
print('M�rkte')
print(round(mean_time - start_time, 1), f'({mean_time - mean_time0:.1f})')
mean_time0 = mean_time
for co in range(NUM_COMPANIES):
    reportWB_write = load_workbook(
        filename = company_dir[co] + report_files[co])
    reportWS_write = reportWB_write['Erl�sRechnung']
    reportWS_write['E10'].value = sales_cur[0, co]
    reportWS_write['E11'].value = sales_cur[1, co]
    reportWS_write['E12'].value = mDec_SOLID_h[8, co, PERIOD+OFFSET]
    reportWS_write['E23'].value = sales_cur[2, co]
    reportWS_write['E24'].value = sales_cur[3, co]
    reportWS_write['L13'].value = mDec_SOLID_h[6, co, PERIOD+OFFSET]
    reportWS_write['L26'].value = mDec_IDEAL_h[6, co, PERIOD+OFFSET]
    reportWS_write = reportWB_write['Personal']
    reportWS_write['D9'].value  = pRes_HR_h[ 1, co, PERIOD+OFFSET]
    reportWS_write['H6'].value  = pRes_HR_h[19:21, co, PERIOD+OFFSET].sum()
    reportWS_write['H7'].value  = pRes_HR_h[13, co, PERIOD+OFFSET]
    reportWS_write['F30'].value = pRes_HR_h[15, co, PERIOD+OFFSET]
    reportWS_write['F31'].value = pRes_HR_h[18, co, PERIOD+OFFSET]
    reportWS_write['F32'].value = pRes_HR_h[14, co, PERIOD+OFFSET]
    reportWS_write = reportWB_write['Anlageverm�gen']
    reportWS_write['D25'].value = mMix_SOLID_h[4, co, PERIOD+OFFSET]-100
    reportWS_write['E25'].value = mMix_SOLID_h[5, co, PERIOD+OFFSET]-100
    reportWS_write['F25'].value = mMix_IDEAL_h[4, co, PERIOD+OFFSET]-100
    reportWS_write['G25'].value = mMix_IDEAL_h[5, co, PERIOD+OFFSET]-100
    reportWS_write = reportWB_write['Vorr�te']
    reportWS_write['C16'].value = -express_SOLID[co]
    reportWS_write['F15'].value = mDec_SOLID_h[6, co, PERIOD+OFFSET]
    reportWS_write['I16'].value = -express_IDEAL[co]
    reportWS_write['L15'].value = mDec_IDEAL_h[6, co, PERIOD+OFFSET]
    reportWB_write.save(filename = company_dir[co] + report_files[co])
    reportWB_write.close()
if (PERIOD == 0):
    reportWB  =  load_workbook(
        filename = company_dir[0] + report_files[0])
    decisionWB = load_workbook(
        filename = company_dir[0] + decision_files_new[0])
    for co in range(1, num_companies0):
        reportWB.save(filename = company_dir[co] + report_files[co])
        decisionWB.save(filename = company_dir[co] + decision_files_new[co])
    del(co)
    reportWB.close()
    decisionWB.close()
    del(reportWB, decisionWB)
del(decision_files, decision_files_new)
del(mDec_SOLID, mDec_IDEAL, mDec_GESAMT,
    pDec_SOLID, pDec_IDEAL, pDec_HR,
    pAll_TA_inv, pAll_TA_sell, pAll_TA_buy, fDec_FIN)
del(sales_cur, express_SOLID, express_IDEAL)
del(type_TA, ndx)
del(reportWB_write, reportWS_write)
mean_time = time.time()
print('')
print('2. Schleife')
print(round(mean_time - start_time, 1), f'({mean_time - mean_time0:.1f})')
mean_time0 = mean_time
with xw.App(visible=False) as app:
    for co in range(num_companies0):
        report_WB = xw.Book(company_dir[co] + report_files[co])
        info_period_WS  = info_period_WB['Branchenbericht']
        report_WS = report_WB.sheets['GuV']
        value_list = report_WS['F5:F14'].value
        value_list = np.around([0 if el is None else el for el in value_list], 3)
        info_period_WS.cell(row= 6, column=4+co).value = value_list[0]
        info_period_WS.cell(row= 7, column=4+co).value = value_list[2]
        info_period_WS.cell(row= 9, column=4+co).value = value_list[4]
        info_period_WS.cell(row=10, column=4+co).value = value_list[5]
        info_period_WS.cell(row=11, column=4+co).value = value_list[6]
        info_period_WS.cell(row=12, column=4+co).value = value_list[7] + value_list[8]
        fRes_COMP_h[0, co, PERIOD+OFFSET] = value_list[9]
        value_list = report_WS['C16:C27'].value
        value_list = np.around([0 if el is None else el for el in value_list], 3)
        info_period_WS.cell(row=14, column=4+co).value = value_list[0]
        info_period_WS.cell(row=15, column=4+co).value = value_list[1]
        info_period_WS.cell(row=16, column=4+co).value = value_list[3] + value_list[4]
        info_period_WS.cell(row=18, column=4+co).value = value_list[10]
        fRes_COMP_h[1, co, PERIOD+OFFSET] = value_list[11]
        value_list = report_WS['K26:K28'].value
        value_list = np.around([0 if el is None else el for el in value_list], 3)
        fRes_COMP_h[2, co, PERIOD+OFFSET] = value_list[0]
        fRes_COMP_h[3, co, PERIOD+OFFSET] = value_list[2]
        report_WS = report_WB.sheets['Bilanz']
        value_list = report_WS['D5:D15'].value
        value_list = np.around([0 if el is None else el for el in value_list], 3)
        info_period_WS.cell(row=23, column=4+co).value = value_list[0]
        info_period_WS.cell(row=24, column=4+co).value = value_list[1]
        info_period_WS.cell(row=25, column=4+co).value = value_list[2]
        info_period_WS.cell(row=26, column=4+co).value = value_list[3]
        info_period_WS.cell(row=28, column=4+co
                            ).value =  value_list[6] + value_list[7]
        info_period_WS.cell(row=29, column=4+co).value = value_list[8]
        info_period_WS.cell(row=30, column=4+co).value = value_list[9]
        info_period_WS.cell(row=31, column=4+co).value = value_list[10]
        fRes_COMP_h[5, co, PERIOD+OFFSET] = value_list[4]
        fDec_FIN_h[7, co, PERIOD+OFFSET] = value_list[10]
        value_list = report_WS['I5:I18'].value
        value_list = np.around([0 if el is None else el for el in value_list], 3)
        info_period_WS.cell(row=35, column=4+co
                            ).value =  value_list[0] + value_list[1]
        info_period_WS.cell(row=36, column=4+co).value = value_list[2]
        info_period_WS.cell(row=37, column=4+co).value = value_list[3]
        info_period_WS.cell(row=39, column=4+co).value =  value_list[6] + value_list[7]
        info_period_WS.cell(row=40, column=4+co).value =  value_list[8] + value_list[8]
        info_period_WS.cell(row=41, column=4+co).value = value_list[10]
        fDec_FIN_h[6, co, PERIOD+OFFSET]  = value_list[10]
        fRes_COMP_h[6, co, PERIOD+OFFSET] = value_list[11]
        fRes_COMP_h[7, co, PERIOD+OFFSET] = value_list[13]
        report_WS = report_WB.sheets['Liquidit�tsRechnung']
        value_list = report_WS['F17'].value
        fRes_COMP_h[4, co, PERIOD+OFFSET] = value_list
        (yield_spread, comp_rating) = mod.rating(
            fRes_COMP_h[4:8, co, PERIOD+OFFSET-1:PERIOD+OFFSET+1])
        fRes_COMP_h[8, co, PERIOD+OFFSET] = yield_spread
        report_WS = report_WB.sheets['Finanzwerte']
        report_WS['C7'].value  = comp_rating
        report_WS['G7'].value  = yield_spread
        info_period_WS.cell(row=64, column=4+co).value =  comp_rating
        info_period_WS.cell(row=65, column=4+co).value =  yield_spread
        report_WS = report_WB.sheets['Personal']
        value_list = report_WS['G9'].value
        pRes_HR_h[5, co, PERIOD+OFFSET] = value_list
        report_WS = report_WB.sheets['Kostenrechnung 2']
        cost_1 = report_WS['H11'].value
        cost_2 = report_WS['H18'].value
        cost_3 = report_WS['I11'].value
        cost_4 = report_WS['I18'].value
        info_period_WS  = info_period_WB['Marktbericht SOLID']
        info_period_WS.cell(row=34, column=4+co).value = cost_1
        info_period_WS.cell(row=35, column=4+co).value = cost_2
        if PERIOD >= MARKET_2:
            info_period_WS  = info_period_WB['Marktbericht IDEAL']
            info_period_WS.cell(row=32, column=4+co).value = cost_3
            info_period_WS.cell(row=33, column=4+co).value = cost_4
        pRes_costs_h[0, co, PERIOD+OFFSET] = cost_1
        pRes_costs_h[1, co, PERIOD+OFFSET] = cost_2
        pRes_costs_h[2, co, PERIOD+OFFSET] = cost_3
        pRes_costs_h[3, co, PERIOD+OFFSET] = cost_4
        report_WB.save()
        info_period_WS = info_period_WB[f'KonjunkturAusblick P{PERIOD+1:0>2d}']
del(app, value_list, yield_spread, comp_rating,
    cost_1, cost_2, cost_3, cost_4)
for co in range(num_companies0):
    info_period_WB.save(filename = company_dir[co] + info_files[co])
info_period_WB.save(filename = GMS_directory + info_period_file)
info_period_WB.close()
del(info_period_WS, info_period_WB)
del(report_files, report_WS, report_WB)
mean_time = time.time()
print('')
print('Speichere Excel-Berichte')
print(round(mean_time - start_time, 1), f'({mean_time - mean_time0:.1f})')
mean_time0 = mean_time
for co in range(num_companies0):
    info_WB  = load_workbook(filename = company_dir[co] + info_files[co])
    if mDec_GESAMT_h[3, co, PERIOD+OFFSET] == 0:
        info_WS = info_WB['Branchenbericht']
        info_WB.remove(info_WS)
    if mDec_SOLID_h[7, co, PERIOD+OFFSET] == 0:
        info_WS = info_WB['Marktbericht SOLID']
        info_WB.remove(info_WS)
    info_WB.save(filename = company_dir[co] + info_files[co])
    info_WB.close()
del(co, num_companies0)
del(info_WB)
np.savez(GMS_directory + SCENARIO_FILE,
          Preisindex  = szenario['Preisindex'],
          Finanzen    = szenario['Finanzen'],
          Zinsen      = szenario['Zinsen'],
          Sondermarkt = szenario['Sondermarkt'],
          MarktVol    = szenario['MarktVol'],
          eQI_SOLID   = szenario['eQI_SOLID'],
          eQI_IDEAL   = szenario['eQI_IDEAL'],
          MengeHStoff = szenario['MengeHStoff'],
          PreisHStoff = szenario['PreisHStoff'],
          MengeMat1   = szenario['MengeMat1'],
          PreisMat1   = szenario['PreisMat1'],
          MengeMat2   = szenario['MengeMat2'],
          PreisMat2   = szenario['PreisMat2'],
          LagerMat    = szenario['LagerMat'],
          LagerFE     = szenario['LagerFE'],
          TransportFE = szenario['TransportFE'],
          BedarfSOLID = szenario['BedarfSOLID'],
          BedarfIDEAL = szenario['BedarfIDEAL'],
          Gehalt      = szenario['Gehalt'],
          LohnNK      = szenario['LohnNK'],
          eMA         = szenario['eMA'],
          AV_GG       = szenario['AV_GG'],
          AnlagenS    = szenario['AnlagenS'],
          AnlagenM    = szenario['AnlagenM'],
          AnlagenL    = szenario['AnlagenL'],
          AV_BGA      = szenario['AV_BGA'])
np.savez(GMS_directory + COMPANY_FILE,
         mDec_SOLID  = mDec_SOLID_h,
         mDec_IDEAL  = mDec_IDEAL_h,
         mDec_GESAMT = mDec_GESAMT_h,
         mMix_SOLID  = mMix_SOLID_h,
         cSat_SOLID  = cSat_SOLID_h,
         mMix_IDEAL  = mMix_IDEAL_h,
         cSat_IDEAL  = cSat_IDEAL_h,
         pDec_SOLID  = pDec_SOLID_h,
         pDec_IDEAL  = pDec_IDEAL_h,
         pRes_TA     = pRes_TA_h,
         pDec_HR     = pDec_HR_h,
         pRes_HR     = pRes_HR_h,
         pRes_costs  = pRes_costs_h,
         fDec_FIN    = fDec_FIN_h,
         fRes_COMP   = fRes_COMP_h)
del(szenario)
del(mDec_SOLID_h, mDec_IDEAL_h, mDec_GESAMT_h,
    mMix_SOLID_h, cSat_SOLID_h, mMix_IDEAL_h, cSat_IDEAL_h,
    pDec_SOLID_h, pDec_IDEAL_h, pRes_TA_h, pDec_HR_h, pRes_HR_h, pRes_costs_h,
    fDec_FIN_h, fRes_COMP_h)
end_time = time.time()
print('')
print('Aktualisiere Informations-Datei')
print(round(end_time - start_time, 1), f'({end_time - mean_time0:.1f})')
print('Gesamt: ', round(end_time - start_time, 1))
del(start_time, mean_time, mean_time0, end_time)
del(MARKET_0, MARKET_1, MARKET_2, MARKET_3, IDEAL_RD, NUM_CELLS, UL_CELLS)
del(info_files, info_period_file)
del(version_dir, scenario_dir, company_dir)